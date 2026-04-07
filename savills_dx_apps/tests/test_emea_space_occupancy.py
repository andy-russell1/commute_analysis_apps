from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from apps.emea_space_occupancy import engine, io, validation
from apps.emea_space_occupancy import visuals as eso_visuals
from apps.lens.core.constants import SAVILLS_COLOR_SEQUENCE


WORKBOOK_PATH = (
    Path(__file__).resolve().parents[1]
    / "apps"
    / "emea_space_occupancy"
    / "assets"
    / "emea_space_occupancy_demo_dataset.xlsx"
)


def _load_validation(path: Path = WORKBOOK_PATH):
    bundle = io.load_workbook_from_bytes(path.read_bytes(), path.name)
    return validation.validate_workbook(bundle)


def _drop_sheet(sheet_name: str) -> bytes:
    workbook = load_workbook(WORKBOOK_PATH)
    del workbook[sheet_name]
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _all_filters() -> dict[str, object]:
    return {
        "region": [],
        "country": [],
        "city": [],
        "site_name": [],
        "site_type": [],
        "building_name": [],
        "business_unit": [],
        "month_range": None,
    }


def test_demo_workbook_validates_and_supports_scenarios():
    result = _load_validation()
    assert result["status"] in {"Ready", "Ready with warnings"}
    assert "portfolio_hierarchy" in result["clean_sheets"]
    assert "scenario_outputs" in result["clean_sheets"]
    assert engine.get_scenario_names(result["clean_sheets"]) == [
        "Base 2026",
        "Consolidation 2027",
        "Growth Focus 2028",
        "RTO Uplift 2027",
    ]


def test_missing_scenario_outputs_is_warning_not_blocking():
    workbook_bytes = _drop_sheet("scenario_outputs")
    bundle = io.load_workbook_from_bytes(workbook_bytes, "missing_outputs.xlsx")
    result = validation.validate_workbook(bundle)
    issues = result["issues"]
    scenario_output_issue = issues[issues["sheet"] == "scenario_outputs"].iloc[0]
    assert scenario_output_issue["severity"] == "Warning"
    assert result["blocking"] is False

    assumptions = engine.build_working_assumptions(result["clean_sheets"], "Base 2026")
    live = engine.compute_live_scenario(
        result["clean_sheets"],
        assumptions,
        _all_filters(),
        workbook_name="missing_outputs.xlsx",
        workbook_hash="hash",
    )
    assert len(live["outputs"]) == 50
    assert live["origin"] == "Live Scenario"


def test_live_scenario_outputs_cover_portfolio_and_actions():
    result = _load_validation()
    assumptions = engine.build_working_assumptions(result["clean_sheets"], "Consolidation 2027")
    live = engine.compute_live_scenario(
        result["clean_sheets"],
        assumptions,
        _all_filters(),
        workbook_name=WORKBOOK_PATH.name,
        workbook_hash="hash",
    )
    assert len(live["outputs"]) == 50
    assert set(live["outputs"]["action_flag"].dropna().unique()) <= {
        "Expand",
        "Consolidate / Release Space",
        "Exit / Merge",
        "Re-stack / Rebalance",
        "Maintain",
    }
    assert live["summary"]["required_seats"] > 0
    assert live["summary"]["scenario_score"] > 0


def test_interpolate_forecast_supports_short_mid_and_extended_horizons():
    row = {
        "current_headcount": 100,
        "forecast_headcount_12m": 124,
        "forecast_headcount_24m": 160,
    }
    assert engine._interpolate_forecast(row, 6) == 112.0
    assert engine._interpolate_forecast(row, 18) == 142.0
    assert engine._interpolate_forecast(row, 30) == 169.0


def test_baseline_surfaces_standard_precedence_and_sources():
    result = _load_validation()
    baseline = engine.build_portfolio_baseline(result["clean_sheets"], _all_filters())
    site_table = baseline["site_table"]
    assert "default_standard_precedence" in site_table.columns
    assert "default_desk_sharing_ratio_target_source" in site_table.columns
    assert site_table["default_standard_precedence"].str.contains("Workstyle weighted mix").all()
    assert site_table["default_desk_sharing_ratio_target_source"].notna().all()


def test_live_scenario_exposes_component_scores_and_explanations():
    result = _load_validation()
    assumptions = engine.build_working_assumptions(result["clean_sheets"], "Growth Focus 2028")
    live = engine.compute_live_scenario(
        result["clean_sheets"],
        assumptions,
        _all_filters(),
        workbook_name=WORKBOOK_PATH.name,
        workbook_hash="hash",
    )
    outputs = live["outputs"]
    assert {"action_reason", "why_this_changed", "capacity_fit_score", "utilisation_fit_score"}.issubset(outputs.columns)
    assert outputs["action_reason"].astype(str).str.len().gt(0).all()
    assert outputs["why_this_changed"].astype(str).str.len().gt(0).all()
    assert live["summary"]["capacity_fit_score"] > 0
    assert live["summary"]["implementation_simplicity_score"] > 0


def test_consolidation_scenario_tracks_transfer_flow_without_losing_headcount():
    result = _load_validation()
    assumptions = engine.build_working_assumptions(result["clean_sheets"], "Consolidation 2027")
    live = engine.compute_live_scenario(
        result["clean_sheets"],
        assumptions,
        _all_filters(),
        workbook_name=WORKBOOK_PATH.name,
        workbook_hash="hash",
    )
    outputs = live["outputs"]
    transfer_out = outputs["transfer_out_headcount"].fillna(0.0).sum()
    transfer_in = outputs["transfer_in_headcount"].fillna(0.0).sum()
    assert transfer_out > 0
    assert transfer_in > 0
    assert abs(transfer_in - transfer_out) < 1e-6


def test_space_plan_includes_allocation_basis_and_confidence():
    result = _load_validation()
    assumptions = engine.build_working_assumptions(result["clean_sheets"], "Base 2026")
    live = engine.compute_live_scenario(
        result["clean_sheets"],
        assumptions,
        _all_filters(),
        workbook_name=WORKBOOK_PATH.name,
        workbook_hash="hash",
    )
    plans = engine.build_space_plan(result["clean_sheets"], live, _all_filters())
    assert {"allocation_basis", "planning_confidence", "action_reason"}.issubset(plans["building_plan"].columns)
    assert {"allocation_basis", "planning_confidence", "why_this_changed"}.issubset(plans["floor_plan"].columns)


def test_snapshot_contains_audit_metadata():
    result = _load_validation()
    assumptions = engine.build_working_assumptions(result["clean_sheets"], "Base 2026")
    live = engine.compute_live_scenario(
        result["clean_sheets"],
        assumptions,
        _all_filters(),
        workbook_name=WORKBOOK_PATH.name,
        workbook_hash="hash",
    )
    snapshot = engine.build_snapshot(
        live,
        active_filters={"region": ["Northern Europe"]},
        origin="Saved Scenario Snapshot",
        scenario_name="Client Test Scenario",
    )
    assert snapshot["scenario_name"] == "Client Test Scenario"
    assert snapshot["origin"] == "Saved Scenario Snapshot"
    assert snapshot["workbook_name"] == WORKBOOK_PATH.name
    assert snapshot["workbook_hash"] == "hash"
    assert snapshot["filters"] == {"region": ["Northern Europe"]}
    assert not snapshot["assumptions_used"].empty
    assert not snapshot["calculated_outputs"].empty


def test_seed_scenarios_backfill_derived_output_columns_for_comparison_views():
    result = _load_validation()
    seed_snapshots = engine.build_seed_snapshots(
        result["clean_sheets"],
        workbook_name=WORKBOOK_PATH.name,
        workbook_hash="hash",
    )

    assert seed_snapshots
    first_outputs = seed_snapshots[0]["calculated_outputs"]
    assert "scenario_score" in first_outputs.columns
    assert "key_risk" in first_outputs.columns

    comparison_table = engine.comparison_site_table(seed_snapshots[:2])
    assert not comparison_table.empty
    assert {"scenario_score", "key_risk", "region"}.issubset(comparison_table.columns)


def test_emea_visuals_use_lens_colour_sequence():
    donut_df = engine.build_portfolio_baseline(_load_validation()["clean_sheets"], _all_filters())["space_mix"].head(4).copy()
    donut_fig = eso_visuals.donut_chart(donut_df, names="space_type", values="area_sqm", title="Area by Space Type")
    assert donut_fig is not None
    assert tuple(donut_fig.layout.colorway)[:4] == tuple(SAVILLS_COLOR_SEQUENCE[:4])

    bar_df = donut_df.rename(columns={"space_type": "Category", "area_sqm": "Value"})
    bar_fig = eso_visuals.bar_chart(bar_df, x="Category", y="Value", title="Test")
    assert bar_fig is not None
    assert tuple(bar_fig.layout.colorway)[:4] == tuple(SAVILLS_COLOR_SEQUENCE[:4])
