from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from apps.lens.core import io


def test_template_has_required_sheets_and_parser_support():
    workbook = Path("apps/lens/assets/LENS.xlsx")
    assert workbook.exists(), "Expected canonical template workbook at apps/lens/assets/LENS.xlsx."

    try:
        wb = load_workbook(workbook, data_only=True)
    except PermissionError as exc:  # pragma: no cover - environment specific file lock
        pytest.skip(f"Workbook locked by another process: {exc}")

    sheet_names = wb.sheetnames
    assert "How_To_Use" in sheet_names
    assert "Criteria Sheet" in sheet_names
    assert "Data Sheet" in sheet_names

    try:
        parsed = io.load_workbook_from_bytes(workbook.read_bytes())
    except PermissionError as exc:  # pragma: no cover - environment specific file lock
        pytest.skip(f"Workbook locked by another process: {exc}")

    assert "criteria" in parsed
    assert "raw_data" in parsed
    assert "city_columns" in parsed


